import streamlit as st
import hashlib
import time
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import base64
import os
import sys
from pathlib import Path
import difflib
import io
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter

st.set_page_config(page_title="Takttool – Montage- & Personalplanung", layout="wide")

# --- Passwortschutz ---
def check_password():
    def hash_password(password: str) -> str:
        return hashlib.sha256(password.encode()).hexdigest()

    # mehrere gültige Passwörter
    valid_hashes = {
        hash_password("Targus2025!"),
        hash_password("Stadler2025!"),
    }

    if "auth_ok" not in st.session_state:
        st.session_state["auth_ok"] = False

    if not st.session_state["auth_ok"]:
        st.markdown("## 🔐 Geschützter Bereich")
        with st.form("login_form"):
            password = st.text_input("Bitte Passwort eingeben", type="password")
            submitted = st.form_submit_button("Einloggen")
            if submitted:
                if hash_password(password) in valid_hashes:
                    st.session_state["auth_ok"] = True
                    st.success("✅ Login erfolgreich – lade App…")
                    st.rerun()
                else:
                    st.error("❌ Falsches Passwort")
                    st.stop()
        st.stop()

check_password()

# --- Styles (Dark) ---
st.markdown("""
    <style>
    html, body, [data-testid="stApp"] { background-color: #1a1a1a; color: #ffffff; font-family: 'Segoe UI', sans-serif; }
    h1, h2, h3 { color: #CC0000; text-align: center; }
    div[data-testid="stFileUploader"] > label { font-size: 0.8rem; padding-bottom: 0.25rem; margin-bottom: 0.25rem; }
    section[data-testid="stFileUploaderDropzone"] { padding: 0.2rem 0.5rem; background-color: #2a2a2a; border: 1px solid #444; border-radius: 6px; text-align: center; }
    div[data-testid="stFileUploader"] { margin-bottom: 0.25rem; }
    .stDataFrameContainer { border-radius: 10px; border: 1px solid #444; }
    div[data-baseweb="tabs"] { margin-top: 1rem; }
    button[data-baseweb="tab"] { font-size: 20px !important; padding: 12px 20px !important; margin: 0 !important; height: auto !important; border-radius: 0 !important; border: none !important; background-color: #2a2a2a !important; color: #ddd !important; transition: background-color 0.3s ease; }
    button[data-baseweb="tab"][aria-selected="true"] { background-color: #CC0000 !important; color: white !important; font-weight: bold; }
    button[data-baseweb="tab"] + button[data-baseweb="tab"] { border-left: 1px solid #1a1a1a; }
    [data-testid="stHeader"] { background: transparent !important; border: none !important; box-shadow: none !important; height: 0px !important; }
    header, .st-emotion-cache-18ni7ap { background: transparent !important; border: none !important; box-shadow: none !important; }
    </style>
""", unsafe_allow_html=True)

# --- Canonicals & Defaults ---
CANONICALS = ["Datum", "Tag", "Takt", "Soll-Zeit", "Qualifikation", "Inhalt", "Bauraum"]
DEFAULT_PLAN_TYPES = ["EW1", "EW2", "MW1", "MW2"]

if "plan_types" not in st.session_state:
    st.session_state["plan_types"] = DEFAULT_PLAN_TYPES.copy()
if "num_wagen" not in st.session_state:
    st.session_state["num_wagen"] = 12

for key in st.session_state["plan_types"]:
    st.session_state.setdefault(f"df_{key}", pd.DataFrame())
    st.session_state.setdefault(f"map_{key}", {})
    st.session_state.setdefault(f"file_{key}", None)            # UploadedFile
    st.session_state.setdefault(f"file_bytes_{key}", None)      # bytes (für Write-Back)
    st.session_state.setdefault(f"file_name_{key}", None)       # Originalname

DEFAULT_HINTS = {
    "Inhalt": ["Baugruppe / Arbeitsgang", "Arbeitsgang", "Inhalt"],
    "Soll-Zeit": ["Std.", "Stunden", "Soll-Zeit"],
    "Bauraum": ["Ebene", "Bauraum"],
    "Datum": ["Datum", "Datum \\nStart (Berechnet)", "Startdatum"],
    "Qualifikation": ["Qualifikation", "Skill"],
    "Tag": ["Tag", "MAP-Tag", "Tag (MAP)"],
    "Takt": ["Takt", "Station", "Taktnummer"]
}

def propose_for(canon, cols):
    for hint in DEFAULT_HINTS.get(canon, []):
        if hint in cols:
            return hint
    m = difflib.get_close_matches(canon, cols, n=1, cutoff=0.6)
    return m[0] if m else None

def _col_as_series(df: pd.DataFrame, name: str):
    if name not in df.columns:
        return pd.Series([np.nan] * len(df), index=df.index)
    data = df.loc[:, df.columns == name]
    if isinstance(data, pd.Series):
        return data
    if data.shape[1] == 1:
        return data.iloc[:, 0]
    def pick_first_valid(row):
        for x in row:
            if pd.notna(x) and str(x).strip() != "":
                return x
        return np.nan
    return data.apply(pick_first_valid, axis=1)

# --- Helfer: Erste Excel-Tabelle (Als Tabelle formatiert) robust einlesen ---
def _read_first_excel_table(uploaded_or_bytes) -> pd.DataFrame:
    """
    Liest bei XLSX den Bereich der ersten 'als Tabelle formatierten' Excel-Tabelle
    und gibt ihn als DataFrame zurück. Erste Zeile im Bereich = Header.
    Wirft ValueError('NO_TABLE'), wenn keine Tabelle existiert.
    """
    if isinstance(uploaded_or_bytes, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(uploaded_or_bytes), data_only=True)
    else:
        wb = load_workbook(uploaded_or_bytes, data_only=True)

    tables = []
    for ws in wb.worksheets:
        for t in ws._tables.values():
            tables.append((ws, t.ref))
    if not tables:
        raise ValueError("NO_TABLE")

    ws, ref_range = tables[0]
    min_col, min_row, max_col, max_row = range_boundaries(ref_range)

    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col,
                             values_only=True))
    if not rows:
        return pd.DataFrame()

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = rows[1:]

    def _make_unique(cols):
        seen = {}
        out = []
        for c in cols:
            c0 = c if c else "Spalte"
            if c0 not in seen:
                seen[c0] = 0
                out.append(c0)
            else:
                seen[c0] += 1
                out.append(f"{c0}_{seen[c0]}")
        return out

    header = _make_unique(header)
    df = pd.DataFrame(data, columns=header)
    return df

# --- Datei einlesen + Mapping anwenden ---
def lade_und_verarbeite_datei_mit_mapping(uploaded_file, mapping_canonical_to_source: dict, file_bytes=None):
    df = pd.DataFrame()
    if uploaded_file is None and file_bytes is None:
        return df
    try:
        if (uploaded_file is not None and uploaded_file.name.lower().endswith(".xlsx")) or (file_bytes is not None):
            try:
                df = _read_first_excel_table(file_bytes if file_bytes is not None else uploaded_file)
            except ValueError as ve:
                if str(ve) == "NO_TABLE":
                    st.error("❌ Bitte den Bereich des MAP als **Tabelle** in Excel formatieren (Einfügen → Tabelle).")
                    return pd.DataFrame()
                else:
                    raise
        else:
            if uploaded_file is not None:
                df = pd.read_csv(uploaded_file)
            else:
                st.error("CSV benötigt eine UploadedFile-Referenz.")
                return pd.DataFrame()

        df.columns = [str(c).strip() for c in df.columns]

        valid_map = {canon: src for canon, src in mapping_canonical_to_source.items() if src in df.columns}
        df = df.rename(columns={src: canon for canon, src in valid_map.items()})

        out = pd.DataFrame(index=df.index)
        for c in CANONICALS:
            out[c] = _col_as_series(df, c)

        for c in CANONICALS:
            if c not in out.columns:
                out[c] = np.nan if c in ["Datum", "Tag", "Takt", "Soll-Zeit"] else ""

        out["Soll-Zeit"] = (out["Soll-Zeit"].astype(str)
                            .str.replace(r"[^\d,\.]", "", regex=True)
                            .str.replace(",", ".", regex=False))
        out["Stunden"] = pd.to_numeric(out["Soll-Zeit"], errors="coerce")
        out = out[out["Stunden"].notna()].copy()

        out["Takt"] = pd.to_numeric(out["Takt"], errors="coerce").fillna(1).astype(int)

        out["Datum"] = pd.to_datetime(out["Datum"], errors="coerce")
        if out["Tag"].isna().all():
            if out["Datum"].notna().any():
                startdatum_ref = out["Datum"].min()
                out["Tag"] = (out["Datum"] - startdatum_ref).dt.days + 1
            else:
                out["Tag"] = 1
        out["Tag"] = pd.to_numeric(out["Tag"], errors="coerce").fillna(1).astype(int)

        if out["Datum"].notna().any():
            out["Start"] = out["Datum"] + pd.to_timedelta(6, unit="h")
            out["Ende"]  = out["Start"] + pd.to_timedelta(out["Stunden"].clip(upper=8), unit="h")
        else:
            out["Start"] = pd.NaT
            out["Ende"]  = pd.NaT

        out["Tag_Takt"] = out["Tag"].astype(str) + "_T" + out["Takt"].astype(str)
        return out.reset_index(drop=True)

    except Exception as e:
        st.error(f"Fehler beim Verarbeiten (Mapping): {e}")
        return pd.DataFrame()

def ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    need = ["Datum", "Tag", "Takt", "Soll-Zeit", "Qualifikation", "Inhalt", "Bauraum",
            "Stunden", "Start", "Ende", "Tag_Takt"]
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame({c: [] for c in need})
    for c in need:
        if c not in df.columns:
            df[c] = np.nan if c in ["Datum","Tag","Takt","Soll-Zeit","Stunden"] else ""
    return df

# --- Montage-Ansicht: gewünschte Spaltenreihenfolge ---
MONTAGE_ORDER_PRIMARY = ["Inhalt", "Tag", "Takt", "Soll-Zeit", "Qualifikation", "Bauraum"]

def order_columns_for_montage(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)
    rest = [c for c in cols if c not in MONTAGE_ORDER_PRIMARY]
    rest_sorted = sorted(rest)
    ordered = [c for c in MONTAGE_ORDER_PRIMARY if c in cols] + rest_sorted
    return df[ordered]

# --- Logo & Titel ---
def zeige_logo_und_titel():
    logo_path = Path("Logo_Targus.png")
    if logo_path.exists():
        logo_bytes = logo_path.read_bytes()
        logo_base64 = base64.b64encode(logo_bytes).decode()
        logo_html = f'''
            <div style="display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; margin-bottom: 1rem;">
                <div style="flex: 1; min-width: 100px; display: flex; align-items: center;">
                    <img src="data:image/png;base64,{logo_base64}" style="max-height: 50px; height: auto; width: auto;">
                </div>
                <div style="flex: 2; text-align: center; min-width: 250px;">
                    <h1 style="margin: 0; font-size: 2rem; color: white;">Takttool: Montage- & Personalplanung</h1>
                </div>
                <div style="flex: 1;"></div>
            </div>
        '''
    else:
        logo_html = '''
            <div style="text-align: center; margin: 1rem;">
                <div style="width:60px; height:60px; background:#ccc; margin: auto;"></div>
                <h1 style="color: white;">Takttool: Montage- & Personalplanung</h1>
            </div>
        '''
    st.markdown(logo_html, unsafe_allow_html=True)

zeige_logo_und_titel()

# --- Plot Helper mit Ø-Linie ---
def bar_with_mean(df_plot, x, y, color, title, height=300):
    fig = px.bar(df_plot, x=x, y=y, color=color, barmode="stack", title=title, height=height)
    try:
        mean_val = df_plot.groupby(x)[y].sum().mean()
        xs = sorted(pd.Series(df_plot[x]).dropna().unique())
        if len(xs) > 0 and pd.notna(mean_val):
            fig.add_trace(go.Scatter(
                x=xs, y=[mean_val] * len(xs), mode="lines", name="Ø pro Tag",
                line=dict(dash="dash", width=2),
                hovertemplate=f"Ø: {mean_val:.2f}<extra></extra>"
            ))
            fig.add_annotation(
                x=xs[-1], y=mean_val, text=f"Ø {mean_val:.1f}",
                showarrow=False, xanchor="left", yanchor="bottom", font=dict(color="#FFFFFF")
            )
    except Exception:
        pass
    fig.update_layout(plot_bgcolor="#1a1a1a", paper_bgcolor="#1a1a1a", font_color="#ffffff", legend_title_text=None)
    return fig

# --- Excel Write-back ---
def write_back_to_excel_table(original_bytes: bytes, edited_df: pd.DataFrame, output_name: str) -> bytes:
    wb = load_workbook(io.BytesIO(original_bytes))
    ws = None
    table_obj = None
    for sheet in wb.worksheets:
        if sheet._tables:
            name, tbl = list(sheet._tables.items())[0]
            ws = sheet
            table_obj = tbl
            break
    if ws is None or table_obj is None:
        raise ValueError("Keine Excel-Tabelle gefunden.")

    min_col, min_row, max_col, max_row = range_boundaries(table_obj.ref)
    header_cells = [ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]
    header = [str(h) if h is not None else "" for h in header_cells]

    df_to_write = edited_df.copy()
    for c in header:
        if c not in df_to_write.columns:
            df_to_write[c] = ""
    df_to_write = df_to_write[header]

    new_rows = len(df_to_write)
    new_max_row = min_row + new_rows
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
    table_obj.ref = new_ref

    for j, colname in enumerate(header, start=min_col):
        ws.cell(row=min_row, column=j, value=colname)

    for r in range(min_row + 1, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c, value=None)

    for i, row in enumerate(df_to_write.itertuples(index=False), start=min_row + 1):
        for j, value in enumerate(row, start=min_col):
            ws.cell(row=i, column=j, value=value)

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.read()

# --- Montage-Tab Renderer ---
def render_montage_tab(df: pd.DataFrame, plan_label: str, slider_key: str, editor_key: str, gantt_key: str, bauraum_prefix: str, quali_prefix: str):
    df = ensure_columns(df)
    st.markdown("#### Zeitraum wählen (nach Tag)")
    tag_vals = pd.to_numeric(df["Tag"], errors="coerce").dropna().astype(int)
    if tag_vals.empty:
        st.warning(f"Keine gültigen Tag-Werte für {plan_label} verfügbar.")
        return
    tag_min, tag_max = int(tag_vals.min()), int(tag_vals.max())
    tag_range = st.slider("Tag auswählen", min_value=tag_min, max_value=tag_max, value=(tag_min, tag_max), key=slider_key)

    df["Tag"] = pd.to_numeric(df["Tag"], errors="coerce").fillna(tag_min).astype(int)
    df_filtered = df[df["Tag"].between(tag_range[0], tag_range[1])].copy()

    if not df_filtered.empty:
        df_filtered = order_columns_for_montage(df_filtered)

    col_table, col_gantt = st.columns([1.25, 1.75])
    with col_table:
        st.markdown("<br><br><br>", unsafe_allow_html=True)
        edited_df = st.data_editor(df_filtered, use_container_width=True, num_rows="dynamic", hide_index=True, key=editor_key)

        excel_buffer = io.BytesIO()
        edited_df.to_excel(excel_buffer, index=False, engine="openpyxl")
        excel_buffer.seek(0)
        st.download_button(
            label="⬇️ Geänderte Datei herunterladen",
            data=excel_buffer,
            file_name=f"Montageplanung_{plan_label}_aktualisiert.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        orig_bytes = st.session_state.get(f"file_bytes_{plan_label}")
        orig_name  = st.session_state.get(f"file_name_{plan_label}") or f"{plan_label}.xlsx"
        if orig_bytes:
            try:
                updated_bytes = write_back_to_excel_table(orig_bytes, edited_df, orig_name)
                st.download_button(
                    label="⬇️ In ursprüngliche Excel-Tabelle zurückschreiben",
                    data=updated_bytes,
                    file_name=(Path(orig_name).stem + "_updated.xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Schreibt in die erste 'Als Tabelle formatiert'-Tabelle der Originaldatei."
                )
            except Exception as wbe:
                st.warning(f"Zurückschreiben in die Tabelle nicht möglich: {wbe}")

        if not edited_df.equals(df_filtered):
            df.update(edited_df)
            st.session_state[f"df_{plan_label}"] = df.copy()
            st.success("Änderungen gespeichert.")

    with col_gantt:
        if not df_filtered.empty:
            if "Start" not in df_filtered.columns or "Ende" not in df_filtered.columns or df_filtered["Start"].isna().all():
                df_filtered["Start"] = pd.to_datetime(df_filtered["Datum"], errors="coerce") + pd.to_timedelta(6, unit="h")
                df_filtered["Ende"]  = df_filtered["Start"] + pd.to_timedelta(df_filtered["Stunden"].clip(upper=8), unit="h")
            fig_gantt = px.timeline(
                df_filtered, x_start="Start", x_end="Ende", y="Inhalt", color="Qualifikation",
                title=f"Ablaufplanung {plan_label}", custom_data=["Tag", "Bauraum", "Stunden"]
            )
            fig_gantt.update_yaxes(autorange="reversed")
            fig_gantt.update_traces(
                hovertemplate=("Tag: %{customdata[0]}<br>Bauraum: %{customdata[1]}<br>Stunden: %{customdata[2]}<br>Inhalt: %{y}<extra></extra>")
            )
            fig_gantt.update_layout(xaxis_title="Datum", yaxis_title=None, plot_bgcolor="#1a1a1a", paper_bgcolor="#1a1a1a", font_color="#ffffff", height=600)
            st.plotly_chart(fig_gantt, use_container_width=True, key=gantt_key)
        else:
            st.info("Keine Daten für Gantt-Diagramm.")

    st.divider()

    if not df_filtered.empty:
        def gruppiere(df_in, group_field):
            return df_in.groupby(["Tag", group_field])["Stunden"].sum().reset_index()

        takte = sorted(pd.to_numeric(df_filtered["Takt"], errors="coerce").dropna().unique())
        bauraum_data = [gruppiere(df_filtered[df_filtered["Takt"] == t], "Bauraum") for t in takte]
        quali_data   = [gruppiere(df_filtered[df_filtered["Takt"] == t], "Qualifikation") for t in takte]
        titel_map    = [f"Takt {int(t)}" for t in takte]

        col_bauraum, col_qualifikation = st.columns(2)
        with col_bauraum:
            st.markdown("### Stunden nach Bauraum")
            for i, df_plot in enumerate(bauraum_data):
                fig = bar_with_mean(df_plot, x="Tag", y="Stunden", color="Bauraum", title=titel_map[i], height=300)
                st.plotly_chart(fig, use_container_width=True, key=f"{bauraum_prefix}_{i}")
        with col_qualifikation:
            st.markdown("### Stunden nach Qualifikation")

            # Normalisierung + feste Farben (Elektriker = Rot #a52019, Mechaniker = Grau #9CA3AF)
            quali_normalize = {
                "elektriker": "Elektriker",
                "elektromonteur": "Elektriker",
                "mechaniker": "Mechaniker",
                "mechaiker": "Mechaniker",
                "mechaikr": "Mechaniker",
                "mech": "Mechaniker",
            }
            fixed_colors = {"Elektriker": "#a52019", "Mechaniker": "#9CA3AF"}
            fallback_palette = list(px.colors.qualitative.Bold)

            for i, df_plot in enumerate(quali_data):
                if df_plot.empty:
                    continue

                _q = df_plot["Qualifikation"].astype(str).str.strip()
                _q_norm = _q.str.lower().map(quali_normalize).fillna(_q)
                df_plot = df_plot.copy()
                df_plot["Qualifikation"] = _q_norm

                fig = bar_with_mean(df_plot, x="Tag", y="Stunden", color="Qualifikation", title=titel_map[i], height=300)

                used = set()
                palette_idx = 0
                for tr in fig.data:
                    name = getattr(tr, "name", None)
                    if name in fixed_colors:
                        tr.update(marker=dict(color=fixed_colors[name]))
                        used.add(fixed_colors[name])
                    else:
                        while palette_idx < len(fallback_palette) and fallback_palette[palette_idx] in used:
                            palette_idx += 1
                        color_choice = fallback_palette[palette_idx % len(fallback_palette)]
                        tr.update(marker=dict(color=color_choice))
                        used.add(color_choice)
                        palette_idx += 1

                st.plotly_chart(fig, use_container_width=True, key=f"{quali_prefix}_{i}")
    else:
        st.info("Keine Daten für Statistiken vorhanden.")

# ==============================
# Tabs (dynamisch)
# ==============================
labels = ["Einrichtung"] + [f"Montageplanung {p}" for p in st.session_state["plan_types"]] + ["Personalplanung", "Export"]
tabs = st.tabs(labels)
tab_setup = tabs[0]
tab_personal = tabs[-2]
tab_export  = tabs[-1]
montage_tabs = tabs[1:-2]

# --- Einrichtung ---
with tab_setup:

    # Plan-Typen anpassen (dynamisch)
    types_csv = st.text_input(
        "Plan-Typen (kommagetrennt):",
        value=", ".join(st.session_state["plan_types"]),
        help="Beispiel: EW1, EW2, MW1, MW2, MW3"
    )
    col_types_left, col_types_right = st.columns([1, 1])
    with col_types_left:
        if st.button("Plan-Typen übernehmen"):
            new_types = [t.strip() for t in types_csv.split(",") if t.strip() != ""]
            if not new_types:
                st.error("Mindestens ein Plan-Typ ist erforderlich.")
            else:
                st.session_state["plan_types"] = list(dict.fromkeys(new_types))
                for key in st.session_state["plan_types"]:
                    st.session_state.setdefault(f"df_{key}", pd.DataFrame())
                    st.session_state.setdefault(f"map_{key}", {})
                    st.session_state.setdefault(f"file_{key}", None)
                    st.session_state.setdefault(f"file_bytes_{key}", None)
                    st.session_state.setdefault(f"file_name_{key}", None)
                st.success("Plan-Typen aktualisiert. Seite wird neu aufgebaut…")
                st.rerun()

    st.markdown("---")
    st.caption("Lade je Plan die Datei hoch und ordne die Quellspalten den erwarteten Spalten zu. Nutze 'Auto-Vorschlag' für eine schnelle Vorbelegung.")

    def mapping_ui(plan_key: str, title: str):
        st.subheader(title)
        up = st.file_uploader(f"Datei für {plan_key} (CSV/XLSX)", type=["csv", "xlsx"], key=f"uploader_{plan_key}")

        if up is not None:
            st.session_state[f"file_{plan_key}"] = up
            st.session_state[f"file_name_{plan_key}"] = up.name
            try:
                st.session_state[f"file_bytes_{plan_key}"] = up.getvalue()
            except Exception:
                raw = up.read()
                st.session_state[f"file_bytes_{plan_key}"] = raw

            try:
                if up.name.lower().endswith(".xlsx"):
                    df_prev = _read_first_excel_table(st.session_state[f"file_bytes_{plan_key}"])
                else:
                    df_prev = pd.read_csv(io.BytesIO(st.session_state[f"file_bytes_{plan_key}"]), nrows=1)
                df_prev.columns = [str(c).strip() for c in df_prev.columns]
            except ValueError as ve:
                if str(ve) == "NO_TABLE":
                    st.error("❌ Bitte den Bereich des MAP als **Tabelle** in Excel formatieren (Einfügen → Tabelle).")
                    df_prev = pd.DataFrame()
                else:
                    st.error(f"Vorschau fehlgeschlagen: {ve}")
                    df_prev = pd.DataFrame()
            except Exception as e:
                st.error(f"Vorschau fehlgeschlagen: {e}")
                df_prev = pd.DataFrame()
        else:
            df_prev = pd.DataFrame()

        cols = list(df_prev.columns)
        if not cols:
            st.info("Bitte eine Datei hochladen, um Spalten zu erkennen.")
            return

        current_map = st.session_state.get(f"map_{plan_key}", {})
        options = ["— nicht zuordnen —"] + cols

        c1, c2 = st.columns([1, 1])
        with c1:
            if st.button("Auto-Vorschlag", key=f"auto_{plan_key}"):
                auto_map = {}
                for canon in CANONICALS:
                    guess = current_map.get(canon) or propose_for(canon, cols)
                    auto_map[canon] = guess
                st.session_state[f"map_{plan_key}"] = auto_map
                current_map = auto_map
        with c2:
            if st.button("Zurücksetzen", key=f"reset_{plan_key}"):
                st.session_state[f"map_{plan_key}"] = {}
                current_map = {}

        new_map = {}
        left, right = st.columns(2)
        with left:
            st.markdown("**Erwartet (Canonical)**")
            for canon in CANONICALS:
                st.markdown(f"- {canon}")
        with right:
            st.markdown("**Zuordnen aus Quelle**")
            for canon in CANONICALS:
                default = current_map.get(canon) or propose_for(canon, cols) or "— nicht zuordnen —"
                sel = st.selectbox(
                    f"{canon}",
                    options,
                    index=options.index(default) if default in options else 0,
                    key=f"map_{plan_key}_{canon}"
                )
                new_map[canon] = None if sel == "— nicht zuordnen —" else sel

        used = [s for s in new_map.values() if s]
        duplicates = {x for x in used if used.count(x) > 1}
        if duplicates:
            st.error("Konflikt: mehrfach zugeordnet → " + ", ".join(sorted(duplicates)))

        if st.button(f"Übernehmen für {plan_key}", key=f"apply_{plan_key}"):
            if st.session_state.get(f"file_{plan_key}") is None:
                st.error("Bitte zuerst eine Datei hochladen.")
            elif duplicates:
                st.error("Bitte Konflikte lösen (jede Quellspalte nur einmal verwenden).")
            else:
                final_map = {canon: src for canon, src in new_map.items() if src}
                df_proc = lade_und_verarbeite_datei_mit_mapping(
                    st.session_state[f"file_{plan_key}"],
                    final_map,
                    file_bytes=st.session_state.get(f"file_bytes_{plan_key}")
                )
                st.session_state[f"map_{plan_key}"] = new_map
                st.session_state[f"df_{plan_key}"] = df_proc
                if not df_proc.empty:
                    st.success(f"✅ {plan_key}: verarbeitet & gespeichert ({len(df_proc)} Zeilen).")
                    st.rerun()
                else:
                    st.warning(f"{plan_key}: Keine verwertbaren Daten nach Verarbeitung.")

    for pt in st.session_state["plan_types"]:
        with st.expander(pt, expanded=(pt == st.session_state["plan_types"][0])):
            mapping_ui(pt, pt)

    st.info("Nach 'Übernehmen' verwenden die Montage-Tabs und die Personalplanung automatisch die gemappten Daten aus diesem Tab.")

# --- Montage-Tabs dynamisch rendern ---
for tab_obj, plan_key in zip(montage_tabs, st.session_state["plan_types"]):
    with tab_obj:
        df_plan = ensure_columns(st.session_state.get(f"df_{plan_key}", pd.DataFrame()))
        render_montage_tab(
            df=df_plan,
            plan_label=plan_key,
            slider_key=f"tag_slider_{plan_key}",
            editor_key=f"data_editor_{plan_key}",
            gantt_key=f"gantt_{plan_key}",
            bauraum_prefix=f"bauraum_plot_{plan_key}",
            quali_prefix=f"quali_plot_{plan_key}"
        )

# --- Hilfsfunktionen für Bewertung / Export ---
def rating_label(ratio: float) -> str:
    if ratio < 0.70:
        return "❌ nicht möglich (<70%)"
    elif ratio >= 0.90:
        return "✅ hervorragend (≥90%)"
    elif ratio >= 0.80:
        return "✅ gut (≥80%)"
    else:
        return "⚠️ ausbaufähig (70–80%)"

def build_pdf_report(pivot_rund: pd.DataFrame, balance_df: pd.DataFrame, overall_text: str) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, title="Personalplanung – Export")

        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph("Personalplanung – Bedarf & Balancing", styles["Title"]))
        story.append(Spacer(1, 12))
        story.append(Paragraph(datetime.now().strftime("%d.%m.%Y %H:%M"), styles["Normal"]))
        story.append(Spacer(1, 18))
        story.append(Paragraph("<b>Gesamtbewertung</b>: " + overall_text, styles["Heading3"]))
        story.append(Spacer(1, 12))

        story.append(Paragraph("Aufgerundete FTE pro Tag & Qualifikation", styles["Heading3"]))
        data1 = [["RelativerTag"] + list(pivot_rund.columns)]
        for idx, row in pivot_rund.iterrows():
            data1.append([idx] + [int(x) for x in row.tolist()])
        t1 = Table(data1)
        t1.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#eeeeee")),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("ALIGN",(1,1),(-1,-1),"RIGHT"),
        ]))
        story.append(t1)
        story.append(Spacer(1, 18))

        story.append(Paragraph("Balancing – Abweichung vom Durchschnitt", styles["Heading3"]))
        data2 = [list(balance_df.columns)]
        for _, r in balance_df.iterrows():
            data2.append([
                int(r["RelativerTag"]),
                f'{r["Total FTE"]:.2f}',
                f'{r["Ø Total FTE"]:.2f}',
                f'{r["Abweichung"]:.1f} %',
                r["Bewertung"]
            ])
        t2 = Table(data2)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#eeeeee")),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("ALIGN",(1,1),(-1,-1),"RIGHT"),
        ]))
        story.append(t2)

        story.append(Spacer(1, 18))
        story.append(Paragraph("Legende: <70% nicht möglich · ≥80% gut · ≥90% hervorragend", styles["Italic"]))

        doc.build(story)
        buffer.seek(0)
        return buffer.read()
    except Exception as e:
        st.warning(f"PDF-Erstellung nicht verfügbar: {e}")
        return b""

# --- Tab: Personalplanung ---
with tab_personal:
    plan_mapping = {}
    for pt in st.session_state["plan_types"]:
        dfp = st.session_state.get(f"df_{pt}", pd.DataFrame())
        if dfp is not None and not dfp.empty:
            plan_mapping[pt] = ensure_columns(dfp)

    if not plan_mapping:
        st.warning("Bitte lade mindestens einen Montageplan im Tab 'Einrichtung' hoch und übernehme das Mapping.")
        st.stop()

    st.markdown("### Parameter")

    st.session_state["num_wagen"] = st.number_input(
        "Anzahl Wagenkästen",
        min_value=1, max_value=50, value=st.session_state["num_wagen"], step=1, help="Gilt für Auswahl & Matrix."
    )

    tag_map_liste = list(range(1, 24))
    max_block = len(tag_map_liste)
    blocklaenge = st.number_input(
        "Anzahl Tage pro Wagen (Blocklänge)",
        min_value=1, max_value=max_block, value=7, step=1, key="blocklaenge_var"
    )

    if "fte_stunden" not in st.session_state:
        st.session_state["fte_stunden"] = 8

    fte_basis = st.number_input(
        "Wieviele Stunden arbeitet ein FTE pro Tag?",
        min_value=1, max_value=24, step=1, key="fte_stunden"
    )

    # NEU: Effizienzgrad berücksichtigt die Produktivität in % (1–100)
    effizienz_grad = st.number_input(
        "Effizienzgrad (%)",
        min_value=1, max_value=100,
        value=st.session_state.get("effizienz_grad", 100),
        step=1,
        key="effizienz_grad",
        help="Produktivitätsfaktor: 100% = voll produktiv, 80% = 1 FTE leistet 80% seiner Stunden."
    )
    effektive_fte_stunden = max(0.01, float(fte_basis) * (float(effizienz_grad) / 100.0))  # Safety gegen 0

    st.caption(f"👉 Effektive FTE-Stunden/Tag: **{effektive_fte_stunden:.2f} h** (Basis {fte_basis} h × Effizienz {effizienz_grad} %)")

    st.markdown("### Auswahl des Montageplans pro Wagenkasten")
    wagen_count = int(st.session_state["num_wagen"])
    wagenkästen = [f"Wagenkasten {i}" for i in range(1, wagen_count + 1)]
    zugewiesene_pläne = {}

    plan_row = st.columns(len(wagenkästen))
    for i, wk in enumerate(wagenkästen):
        zugewiesene_pläne[wk] = plan_row[i].selectbox(
            f"{wk}",
            options=list(plan_mapping.keys()),
            key=f"plan_select_{wk}"
        )

    # Checkbox-Matrix im Form (kein Rerun pro Klick)
    with st.form("belegung_form", clear_on_submit=False):
        st.markdown("### Belegung der MAP-Tage über Checkbox-Matrix")
        header_cols = st.columns([1] + [1] * len(wagenkästen))
        header_cols[0].markdown("**MAP-Tag**")
        for i, wk in enumerate(wagenkästen):
            header_cols[i + 1].markdown(f"**{wk}**")

        for tag_map in tag_map_liste:
            cols = st.columns([1] + [1] * len(wagenkästen))
            cols[0].markdown(f"{tag_map}")
            for wk_idx, wk in enumerate(wagenkästen):
                key = f"wk{wk_idx}_tag{tag_map}"
                current_value = st.session_state.get(key, False)
                cols[wk_idx + 1].checkbox("", value=current_value, key=key)

        btn_autofill = st.form_submit_button("Block aus erstem Häkchen füllen")
        btn_berechne = st.form_submit_button("Berechne Personalbedarf")

    # Autofill
    if btn_autofill:
        for wk_idx, wk in enumerate(wagenkästen):
            selected = [t for t in tag_map_liste if st.session_state.get(f"wk{wk_idx}_tag{t}", False)]
            if not selected:
                continue
            start = min(selected)
            end = min(tag_map_liste[-1], start + int(blocklaenge) - 1)
            for t in tag_map_liste:
                st.session_state[f"wk{wk_idx}_tag{t}"] = False
            for t in range(start, end + 1):
                st.session_state[f"wk{wk_idx}_tag{t}"] = True
        st.rerun()

    # Berechnung
    if btn_berechne:
        zuordnung = {tag_map: [] for tag_map in tag_map_liste}
        for wk_idx, wk in enumerate(wagenkästen):
            for tag_map in tag_map_liste:
                if st.session_state.get(f"wk{wk_idx}_tag{tag_map}", False):
                    zuordnung[tag_map].append((wk_idx, wk))

        # Sanity 1
        fehler_wagen = []
        belegung_pro_wagen = {wk: 0 for wk in wagenkästen}
        for tag_map, einträge in zuordnung.items():
            for _, wk in einträge:
                belegung_pro_wagen[wk] += 1
        for wk, count in belegung_pro_wagen.items():
            if count != 0 and count != int(blocklaenge):
                fehler_wagen.append((wk, count))
        if fehler_wagen:
            fehltext = ", ".join([f"{wk} ({anzahl})" for wk, anzahl in fehler_wagen])
            st.error(f"Fehler: Die folgenden Wagenkästen haben nicht exakt {int(blocklaenge)} Häkchen (oder null): {fehltext}")
            st.stop()

        # Sanity 2
        tag_sets = {name: set(pd.to_numeric(df_src["Tag"], errors="coerce").dropna().astype(int).unique())
                    for name, df_src in plan_mapping.items() if df_src is not None}

        belegte_tage = {
            wk: sorted([tag for tag, einträge in zuordnung.items() if any(w == wk for _, w in einträge)])
            for wk in wagenkästen
        }

        check_rows, kontinuitäts_fehler, not_in_plan_fehler, mapping_rows = [], [], [], []

        for wk in wagenkästen:
            tags = belegte_tage[wk]
            if not tags:
                continue

            plan_name = zugewiesene_pläne[wk]
            tags_sorted = sorted(tags)
            anzahl = len(tags_sorted)
            ist_kontigu = (anzahl == int(blocklaenge)) and (tags_sorted == list(range(tags_sorted[0], tags_sorted[0] + int(blocklaenge))))
            fehlende = sorted([t for t in tags_sorted if t not in tag_sets.get(plan_name, set())])

            bereich_txt = f"{tags_sorted[0]}–{tags_sorted[-1]}" if anzahl > 0 else "-"
            check_rows.append({
                "Wagenkasten": wk,
                "Plan": plan_name,
                "Ausgewählter Bereich": bereich_txt,
                "Anz. Tage": anzahl,
                "Kontiguität OK": "✅" if ist_kontigu else "❌",
                "Fehlende Plan-Tage": ", ".join(map(str, fehlende)) if fehlende else ""
            })

            if not ist_kontigu:
                kontinuitäts_fehler.append(f"{wk} ({bereich_txt})")
            if fehlende:
                not_in_plan_fehler.append(f"{wk} → {plan_name}: {', '.join(map(str, fehlende))}")

            if ist_kontigu and not fehlende:
                for i_rel, t in enumerate(tags_sorted, start=1):
                    mapping_rows.append({
                        "Wagenkasten": wk,
                        "Plan": plan_name,
                        "RelativerTag": i_rel,
                        "Plan-Tag": t
                    })

        st.markdown("#### Sanity-Check Übersicht")
        st.dataframe(pd.DataFrame(check_rows))

        if kontinuitäts_fehler:
            st.error("Nicht zusammenhängende Bereiche gewählt bei: " + ", ".join(kontinuitäts_fehler))
        if not_in_plan_fehler:
            st.error("Ausgewählte Plan-Tage existieren nicht im zugewiesenen Plan:\n- " + "\n- ".join(not_in_plan_fehler))
        if kontinuitäts_fehler or not_in_plan_fehler:
            st.stop()

        if not mapping_rows:
            st.info("Keine validen Zuordnungen gefunden.")
            st.stop()

        df_align = pd.DataFrame(mapping_rows)
        with st.expander("Sanity-Check: Zuordnung (Wagenkasten → RelativerTag → Plan-Tag)", expanded=False):
            st.dataframe(df_align.sort_values(["RelativerTag", "Wagenkasten"]))

        # Daten & Aggregationen
        df_parts = []
        for _, row in df_align.iterrows():
            plan_name = row["Plan"]
            rtag = int(row["RelativerTag"])
            ptag = int(row["Plan-Tag"])
            wk = row["Wagenkasten"]
            df_src = plan_mapping[plan_name].copy()
            df_src["Tag"] = pd.to_numeric(df_src["Tag"], errors="coerce").astype("Int64")
            part = df_src[df_src["Tag"] == ptag].copy()
            if part.empty:
                continue
            part["RelativerTag"] = rtag
            part["Wagenkasten"] = wk
            part["Qualifikation"] = part["Qualifikation"].fillna("Unbekannt")
            df_parts.append(part)

        if not df_parts:
            st.info("Keine Aufgaben für die gewählte Planung gefunden.")
            st.stop()

        df_gesamt = pd.concat(df_parts, ignore_index=True)

        with st.expander("Sanity-Check: Beitrag je Wagenkasten×RelativerTag (Stunden gesamt)", expanded=False):
            pivot = (df_gesamt.groupby(["Wagenkasten", "RelativerTag"])["Stunden"]
                     .sum()
                     .unstack(fill_value=0)
                     .sort_index(axis=1))
            st.dataframe(pivot)

        # --- Diagramme mit Ø-Linie ---
        st.markdown("### Stundenbedarf pro Relativtag")
        df_plot = df_gesamt.groupby(["RelativerTag", "Qualifikation"])["Stunden"].sum().reset_index()

        quali_normalize = {
            "elektriker": "Elektriker",
            "elektromonteur": "Elektriker",
            "elektrik": "Elektriker",
            "mechaniker": "Mechaniker",
            "mechaiker": "Mechaniker",
            "mechaikr": "Mechaniker",
            "mech": "Mechaniker",
        }
        fixed_colors = {"Elektriker": "#a52019", "Mechaniker": "#9CA3AF"}
        fallback_palette = list(px.colors.qualitative.Bold)

        _q = df_plot["Qualifikation"].astype(str).str.strip()
        _q_norm = _q.str.lower().map(quali_normalize).fillna(_q)
        df_plot = df_plot.copy()
        df_plot["Qualifikation"] = _q_norm

        fig_stunden = bar_with_mean(df_plot, x="RelativerTag", y="Stunden", color="Qualifikation",
                                    title="Stundenbedarf pro Relativtag (parallel ausgerichtet)")
        used = set(); palette_idx = 0
        for tr in fig_stunden.data:
            name = getattr(tr, "name", None)
            if name in fixed_colors:
                tr.update(marker=dict(color=fixed_colors[name]))
                used.add(fixed_colors[name])
            else:
                while palette_idx < len(fallback_palette) and fallback_palette[palette_idx] in used:
                    palette_idx += 1
                color_choice = fallback_palette[palette_idx % len(fallback_palette)]
                tr.update(marker=dict(color=color_choice))
                used.add(color_choice)
                palette_idx += 1

        st.plotly_chart(fig_stunden, use_container_width=True)

        st.markdown("### FTE-Bedarf pro Relativtag")
        df_fte = df_plot.copy()
        # HIER: Effizienz wird berücksichtigt
        df_fte["FTE"] = df_fte["Stunden"] / effektive_fte_stunden

        fig_fte = bar_with_mean(df_fte, x="RelativerTag", y="FTE", color="Qualifikation",
                                title="FTE pro Relativtag (mit Ø-Linie, Effizienz berücksichtigt)")
        used = set(); palette_idx = 0
        for tr in fig_fte.data:
            name = getattr(tr, "name", None)
            if name in fixed_colors:
                tr.update(marker=dict(color=fixed_colors[name]))
                used.add(fixed_colors[name])
            else:
                while palette_idx < len(fallback_palette) and fallback_palette[palette_idx] in used:
                    palette_idx += 1
                color_choice = fallback_palette[palette_idx % len(fallback_palette)]
                tr.update(marker=dict(color=color_choice))
                used.add(color_choice)
                palette_idx += 1

        st.plotly_chart(fig_fte, use_container_width=True)

        st.markdown("### Aufgerundete FTE je Relativtag & Qualifikation")
        df_rund = df_fte.copy()
        df_rund["Aufgerundete FTE"] = np.ceil(df_rund["FTE"])
        df_rund = df_rund[["RelativerTag", "Qualifikation", "Aufgerundete FTE"]]
        st.dataframe(df_rund)

        # --- Balancing (Abweichung vom Durchschnitt pro Tag) ---
        day_total_fte = df_fte.groupby("RelativerTag")["FTE"].sum().reset_index(name="Total FTE")
        mean_total = day_total_fte["Total FTE"].mean() if not day_total_fte.empty else 0.0
        day_total_fte["Ø Total FTE"] = mean_total
        day_total_fte["Abweichung"]  = np.where(mean_total > 0, (day_total_fte["Total FTE"] / mean_total) * 100.0, 0.0)
        day_total_fte["Bewertung"]   = day_total_fte["Abweichung"].apply(rating_label)

        min_ratio = day_total_fte["Abweichung"].min() / 100.0 if not day_total_fte.empty else 0.0
        overall_text = rating_label(min_ratio)

        st.markdown("### Balancing – Abweichung vom Durchschnitt pro Tag")
        st.dataframe(day_total_fte.rename(columns={
            "RelativerTag": "RelativerTag",
            "Total FTE": "Total FTE",
            "Ø Total FTE": "Ø Total FTE",
            "Abweichung": "Abweichung (%)",
            "Bewertung": "Bewertung"
        }))

        st.info(f"**Gesamtbewertung Balancing:** {overall_text}  \n"
                "Schwellen: <70 % nicht möglich · ≥80 % gut · ≥90 % hervorragend")

        # --- Ergebnisse für Export merken ---
        st.session_state["pp_df_gesamt"]   = df_gesamt
        st.session_state["pp_df_plot"]     = df_plot
        st.session_state["pp_df_fte"]      = df_fte
        st.session_state["pp_df_rund"]     = df_rund
        st.session_state["pp_balance_df"]  = day_total_fte
        st.session_state["pp_overall_txt"] = overall_text

# --- Tab: Export ---
with tab_export:
    st.markdown("## Export")

    if "pp_df_rund" not in st.session_state or "pp_df_fte" not in st.session_state:
        st.warning("Bitte zuerst im Tab **Personalplanung** auf „Berechne Personalbedarf“ klicken.")
        st.stop()

    df_rund = st.session_state["pp_df_rund"].copy()
    df_fte  = st.session_state["pp_df_fte"].copy()
    balance_df = st.session_state["pp_balance_df"].copy()
    overall_text = st.session_state["pp_overall_txt"]

    # Pivot: Aufgerundete FTE pro Tag & Qualifikation
    pivot_rund = df_rund.pivot_table(index="RelativerTag", columns="Qualifikation",
                                     values="Aufgerundete FTE", aggfunc="sum", fill_value=0).sort_index()
    st.markdown("### Transformierte Tabelle – Bedarf (aufgerundete FTE)")
    st.dataframe(pivot_rund)

    # Optional: Gesamt-FTE mit Ø-Linie
    st.markdown("### Total-FTE pro Tag (mit Ø-Linie)")
    df_total = df_fte.groupby("RelativerTag", as_index=False)["FTE"].sum()
    df_total["Kategorie"] = "Total"
    fig_total = bar_with_mean(df_total.rename(columns={"FTE": "Wert"}),
                              x="RelativerTag", y="Wert", color="Kategorie",
                              title="Total-FTE pro Relativtag (mit Ø-Linie)", height=350)
    st.plotly_chart(fig_total, use_container_width=True)

    # Download: Excel
    excel_buf = io.BytesIO()
    with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
        pivot_rund.to_excel(writer, sheet_name="FTE_aufgerundet")
        st.session_state["pp_df_fte"].to_excel(writer, sheet_name="FTE_detail", index=False)
        st.session_state["pp_df_plot"].to_excel(writer, sheet_name="Stunden_detail", index=False)
        balance_df.to_excel(writer, sheet_name="Balancing", index=False)
    excel_buf.seek(0)
    st.download_button("⬇️ Export als Excel", data=excel_buf,
                       file_name="Personalplanung_Export.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Download: PDF (wenn möglich)
    pdf_bytes = build_pdf_report(pivot_rund, balance_df.rename(columns={"Abweichung": "Abweichung", "Bewertung": "Bewertung"}), overall_text)
    if pdf_bytes:
        st.download_button("⬇️ Export als PDF", data=pdf_bytes, file_name="Personalplanung_Export.pdf", mime="application/pdf")
    else:
        st.caption("Für PDF-Export bitte das Python-Paket **reportlab** installieren.")

# --- Footer / Info für .exe-Nutzung ---
st.markdown("""---""")
st.markdown(
    """
    <div style='text-align: center; font-size: 0.9rem; color: #888;'>
        Entwickelt für Stadlerrail | Urheber: Targus Management Consulting <br>
        For Support contact louis.becker@targusmc.de
    </div>
    """,
    unsafe_allow_html=True
)

import webbrowser
if __name__ == "__main__" and getattr(sys, 'frozen', False):
    time.sleep(2)
    try:
        webbrowser.open("http://localhost:8501")
    except Exception as e:
        print(f"Fehler beim Öffnen des Browsers: {e}")
